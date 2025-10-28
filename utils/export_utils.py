"""
Export Utilities
Export data to Excel and PDF formats
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, A3, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import os

logger = logging.getLogger(__name__)


class ExportUtils:
    """Data export utilities"""
    
    @staticmethod
    def export_to_excel(data: Dict, file_path: str) -> bool:
        """Export data to Excel file with professional formatting"""
        try:
            report_type = data.get('type', 'generic')
            title = data.get('title', 'Rapor')
            records = data.get('data', [])
            
            if not records:
                logger.warning("No data to export")
                return False
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Sınav Programı"
            
            schedule_by_datetime = defaultdict(list)
            for sinav in records:
                if isinstance(sinav.get('tarih_saat'), str):
                    tarih = datetime.fromisoformat(sinav['tarih_saat'])
                else:
                    tarih = sinav['tarih_saat']
                datetime_key = (tarih.date(), tarih.time())
                schedule_by_datetime[datetime_key].append(sinav)
            
            sorted_datetimes = sorted(schedule_by_datetime.keys())
            
            # Get department and exam type info
            bolum_adi = data.get('bolum_adi', '')
            sinav_tipi = data.get('sinav_tipi', '')
            
            # Create title with department and exam type
            full_title = f"{bolum_adi} {sinav_tipi} {title}".upper()
            
            ws.merge_cells('A1:E1')
            title_cell = ws['A1']
            title_cell.value = full_title
            title_cell.font = Font(size=14, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 25
            
            # Add THICK BLACK border below title
            thick_black_bottom = Border(bottom=Side(style='thick', color='000000'))
            for col in range(1, 6):
                cell = ws.cell(row=1, column=col)
                cell.border = Border(
                    top=cell.border.top,
                    left=cell.border.left,
                    right=cell.border.right,
                    bottom=Side(style='thick', color='000000')
                )
            
            headers = ['Tarih', 'Sınav Saati', 'Ders Adı', 'Öğretim Elemanı', 'Derslik']
            ws.append(headers)
            
            header_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in ws[2]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            ws.row_dimensions[2].height = 20
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            current_row = 3
            date_merge_start = {}
            
            # Group exams by (date, time, ders_adi) to merge classrooms
            exam_groups = defaultdict(lambda: {'exams': [], 'derslikler': []})
            for datetime_key in sorted_datetimes:
                date_obj, time_obj = datetime_key
                exams_at_this_time = schedule_by_datetime[datetime_key]
                
                for exam in exams_at_this_time:
                    ders_adi = exam.get('ders_adi', '')
                    key = (date_obj, time_obj, ders_adi)
                    exam_groups[key]['exams'].append(exam)
                    derslik = exam.get('derslikler', exam.get('derslik_adi', ''))
                    if derslik and derslik not in exam_groups[key]['derslikler']:
                        exam_groups[key]['derslikler'].append(derslik)
            
            # Sort by datetime and course name
            sorted_groups = sorted(exam_groups.items(), key=lambda x: (x[0][0], x[0][1], x[0][2]))
            
            for (date_obj, time_obj, ders_adi), group_data in sorted_groups:
                date_str = date_obj.strftime('%d.%m.%Y')
                time_str = time_obj.strftime('%H.%M')
                
                if date_str not in date_merge_start:
                    date_merge_start[date_str] = current_row
                
                # Get first exam info for common fields
                exam = group_data['exams'][0]
                ogretim_elemani = exam.get('ogretim_elemani', '')
                
                # Merge all classrooms into one field
                derslikler_str = '-'.join(group_data['derslikler']) if group_data['derslikler'] else ''
                
                ws.append([date_str, time_str, ders_adi, ogretim_elemani, derslikler_str])
                
                for col in range(1, 6):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(
                        horizontal='center' if col in [1, 2, 5] else 'left',
                        vertical='center',
                        wrap_text=True
                    )
                
                current_row += 1
            
            for date_str, start_row in date_merge_start.items():
                end_row = start_row
                for row in range(start_row, current_row):
                    if ws.cell(row=row, column=1).value == date_str:
                        end_row = row
                
                if end_row > start_row:
                    ws.merge_cells(f'A{start_row}:A{end_row}')
                    merged_cell = ws.cell(row=start_row, column=1)
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                    merged_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    merged_cell.font = Font(bold=True)
            
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 27
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 22

            # Add THICK BLACK borders around entire table
            thick_border = Side(style='thick', color='000000')

            for row in range(1, current_row):
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    current_border = cell.border

                    # Determine which borders should be thick
                    top = thick_border if row == 1 else current_border.top
                    bottom = thick_border if row == current_row - 1 else current_border.bottom
                    left = thick_border if col == 1 else current_border.left
                    right = thick_border if col == 5 else current_border.right

                    cell.border = Border(top=top, bottom=bottom, left=left, right=right)

            wb.save(file_path)
            logger.info(f"Excel exported: {file_path}")
            return True

        except Exception as e:
            logger.error(f"Excel export error: {e}", exc_info=True)
            return False

    @staticmethod
    def export_to_pdf(data: Dict, file_path: str) -> bool:
        """Export data to PDF file with Turkish character support and merged date cells using ReportLab"""
        try:
            report_type = data.get('type', 'sinav_takvimi')
            title = data.get('title', 'Rapor')
            
            # Handle different report types
            if report_type == 'oturma_plani':
                return ExportUtils._export_seating_plan_pdf(data, file_path)
            elif report_type == 'oturma_liste':
                return ExportUtils._export_seating_list_pdf(data, file_path)
            
            # Default: exam schedule
            records = data.get('data', [])
            options = data.get('options', {})

            if not records:
                logger.warning("No data to export")
                return False

            # Register Turkish font - try Arial first (Windows), then DejaVu
            font_registered = False
            font_name = 'TurkishFont'
            font_name_bold = 'TurkishFont-Bold'

            try:
                # Try Arial first (Windows default, supports Turkish)
                arial_paths = {
                    'regular': [
                        'C:/Windows/Fonts/arial.ttf',
                        'C:/Windows/Fonts/Arial.ttf',
                    ],
                    'bold': [
                        'C:/Windows/Fonts/arialbd.ttf',
                        'C:/Windows/Fonts/Arialbd.ttf',
                    ]
                }

                # Try to register Arial
                for font_path in arial_paths['regular']:
                    if os.path.exists(font_path):
                        pdfmetrics.registerFont(TTFont(font_name, font_path))
                        font_registered = True
                        logger.info(f"Registered Arial font from: {font_path}")
                        break

                if font_registered:
                    for font_path in arial_paths['bold']:
                        if os.path.exists(font_path):
                            pdfmetrics.registerFont(TTFont(font_name_bold, font_path))
                            logger.info(f"Registered Arial Bold font from: {font_path}")
                            break

                # If Arial not found, try DejaVu
                if not font_registered:
                    dejavu_paths = {
                        'regular': [
                            'C:/Windows/Fonts/DejaVuSans.ttf',
                            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                        ],
                        'bold': [
                            'C:/Windows/Fonts/DejaVuSans-Bold.ttf',
                            '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
                        ]
                    }

                    for font_path in dejavu_paths['regular']:
                        if os.path.exists(font_path):
                            pdfmetrics.registerFont(TTFont(font_name, font_path))
                            font_registered = True
                            logger.info(f"Registered DejaVu font from: {font_path}")
                            break

                    if font_registered:
                        for font_path in dejavu_paths['bold']:
                            if os.path.exists(font_path):
                                pdfmetrics.registerFont(TTFont(font_name_bold, font_path))
                                logger.info(f"Registered DejaVu Bold font from: {font_path}")
                                break

                if not font_registered:
                    logger.warning("No Turkish-compatible font found, using Helvetica")
            except Exception as e:
                logger.error(f"Font registration failed: {e}")
                font_registered = False

            # Group by date and time
            schedule_by_datetime = defaultdict(list)
            for sinav in records:
                if isinstance(sinav.get('tarih_saat'), str):
                    tarih = datetime.fromisoformat(sinav['tarih_saat'])
                else:
                    tarih = sinav['tarih_saat']
                datetime_key = (tarih.date(), tarih.time())
                schedule_by_datetime[datetime_key].append(sinav)

            sorted_datetimes = sorted(schedule_by_datetime.keys())

            # Group exams by (date, time, ders_adi) to merge classrooms
            exam_groups_pdf = defaultdict(lambda: {'exams': [], 'derslikler': []})
            for datetime_key in sorted_datetimes:
                date_obj, time_obj = datetime_key
                exams_at_this_time = schedule_by_datetime[datetime_key]
                
                for exam in exams_at_this_time:
                    ders_adi = exam.get('ders_adi', '')
                    key = (date_obj, time_obj, ders_adi)
                    exam_groups_pdf[key]['exams'].append(exam)
                    derslik = exam.get('derslikler', exam.get('derslik_adi', ''))
                    if derslik and derslik not in exam_groups_pdf[key]['derslikler']:
                        exam_groups_pdf[key]['derslikler'].append(derslik)
            
            # Sort by datetime and course name
            sorted_groups_pdf = sorted(exam_groups_pdf.items(), key=lambda x: (x[0][0], x[0][1], x[0][2]))
            
            # Build table data with row span info for both date and time
            table_data = []
            table_data.append(['Tarih', 'Sınav Saati', 'Ders Adı', 'Öğretim Elemanı', 'Derslik'])

            date_spans = []  # (start_row, end_row) for date merging
            time_spans = []  # (start_row, end_row) for time merging
            current_date = None
            current_time = None
            date_start_row = 1  # Start after header
            time_start_row = 1

            for (date_obj, time_obj, ders_adi), group_data in sorted_groups_pdf:
                date_str = date_obj.strftime('%d.%m.%Y')
                time_str = time_obj.strftime('%H:%M')

                # Track date changes
                if current_date != date_str:
                    if current_date is not None:
                        date_spans.append((date_start_row, len(table_data) - 1))
                    current_date = date_str
                    date_start_row = len(table_data)

                # Track time changes (for same date)
                datetime_str = f"{date_str}_{time_str}"
                if current_time != datetime_str:
                    if current_time is not None and time_start_row < len(table_data):
                        time_spans.append((time_start_row, len(table_data) - 1))
                    current_time = datetime_str
                    time_start_row = len(table_data)

                # Get first exam info for common fields
                exam = group_data['exams'][0]
                ogretim_elemani = exam.get('ogretim_elemani', '')
                
                # Merge all classrooms into one field
                derslikler_str = '-'.join(group_data['derslikler']) if group_data['derslikler'] else ''

                table_data.append([date_str, time_str, ders_adi, ogretim_elemani, derslikler_str])

            # Add last spans
            if current_date is not None:
                date_spans.append((date_start_row, len(table_data) - 1))
            if current_time is not None and time_start_row < len(table_data):
                time_spans.append((time_start_row, len(table_data) - 1))

            # Create PDF with CUSTOM EXTENDED landscape page - EXTRA LONG for single page
            # Custom page size: A4 width (29.7cm) x EXTENDED height (35cm) - landscape orientation
            custom_pagesize = (35*cm, 29.7*cm)  # (width, height) in landscape

            doc = SimpleDocTemplate(
                file_path,
                pagesize=custom_pagesize,
                topMargin=0.5*cm,
                bottomMargin=0.5*cm,
                leftMargin=0.8*cm,
                rightMargin=0.8*cm
            )

            elements = []

            # Get department and exam type info
            bolum_adi = data.get('bolum_adi', '')
            sinav_tipi = data.get('sinav_tipi', '')

            # Create title with department and exam type
            full_title = f"{bolum_adi} {sinav_tipi} {title}".upper()

            # Title - MINIMAL with BLACK text
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=getSampleStyleSheet()['Heading1'],
                fontSize=11,
                textColor=colors.black,
                spaceAfter=2,
                alignment=TA_CENTER,
                fontName=font_name_bold if font_registered else 'Helvetica-Bold'
            )

            elements.append(Paragraph(full_title, title_style))
            elements.append(Spacer(1, 0.15*cm))

            # Date info - MINIMAL
            date_style = ParagraphStyle(
                'DateStyle',
                parent=getSampleStyleSheet()['Normal'],
                fontSize=6.5,
                alignment=TA_CENTER,
                fontName=font_name if font_registered else 'Helvetica'
            )
            date_text = f"Olusturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            elements.append(Paragraph(date_text, date_style))
            elements.append(Spacer(1, 0.15*cm))

            # Create table with adjusted column widths for EXTENDED landscape page
            # Total width: ~33.4cm for extended page (35cm - 1.6cm margins)
            table = Table(table_data, colWidths=[2.8*cm, 2.8*cm, 16*cm, 8.5*cm, 3.3*cm],
                         repeatRows=1)  # Repeat header on each page

            # Base table style - COMPACT but readable with THICK outer borders
            table_style = TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E67E22')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_name_bold if font_registered else 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                ('TOPPADDING', (0, 0), (-1, 0), 5),

                # All cells - COMPACT PADDING for maximum rows
                ('FONTNAME', (0, 1), (-1, -1), font_name if font_registered else 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7.5),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 1), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 2),

                # Alignment
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Tarih
                ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Saat
                ('ALIGN', (2, 1), (2, -1), 'LEFT'),    # Ders Adı
                ('ALIGN', (3, 1), (3, -1), 'LEFT'),    # Öğretim Elemanı
                ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Derslik

                # Alternating row colors
                ('ROWBACKGROUNDS', (1, 1), (-1, -1), [colors.white, colors.HexColor('#F8F8F8')]),

                # THICK BLACK outer borders for entire table
                ('BOX', (0, 0), (-1, -1), 2.5, colors.black),
            ])

            # Add date cell merging (SPAN) with THICK BLACK borders for separation
            for start_row, end_row in date_spans:
                if end_row > start_row:
                    table_style.add('SPAN', (0, start_row), (0, end_row))
                    table_style.add('BACKGROUND', (0, start_row), (0, end_row), colors.HexColor('#F0F0F0'))
                    table_style.add('FONTNAME', (0, start_row), (0, end_row),
                                  font_name_bold if font_registered else 'Helvetica-Bold')
                    # Add THICK BLACK borders around date groups for clear separation
                    table_style.add('LINEABOVE', (0, start_row), (-1, start_row), 2.5, colors.black)
                    table_style.add('LINEBELOW', (0, end_row), (-1, end_row), 2.5, colors.black)

            # Add time cell merging (SPAN)
            for start_row, end_row in time_spans:
                if end_row > start_row:
                    table_style.add('SPAN', (1, start_row), (1, end_row))
                    table_style.add('BACKGROUND', (1, start_row), (1, end_row), colors.HexColor('#F8F8F8'))
                    table_style.add('FONTNAME', (1, start_row), (1, end_row),
                                  font_name_bold if font_registered else 'Helvetica-Bold')

            table.setStyle(table_style)
            elements.append(table)

            # Build PDF
            doc.build(elements)

            logger.info(f"PDF exported: {file_path}")
            return True

        except Exception as e:
            logger.error(f"PDF export error: {e}", exc_info=True)
            return False

    @staticmethod
    def _export_seating_plan_pdf(data: Dict, file_path: str) -> bool:
        """Export seating plan with visual classroom layout to PDF"""
        try:
            from reportlab.graphics import shapes
            from reportlab.graphics.charts.textlabels import Label
            
            exam_info = data.get('exam_info', {})
            seating_data = data.get('seating_data', {})
            classrooms = data.get('classrooms', [])
            title = data.get('title', 'Oturma Planı')
            
            if not seating_data or not classrooms:
                logger.warning("No seating data or classrooms to export")
                return False
            
            # Use A4 landscape for better printing
            doc = SimpleDocTemplate(
                file_path, 
                pagesize=landscape(A4),
                topMargin=0.5*cm,
                bottomMargin=0.5*cm,
                leftMargin=0.5*cm,
                rightMargin=0.5*cm
            )
            elements = []
            
            # Register Turkish font
            font_registered = False
            font_name = 'TurkishFont'
            try:
                arial_path = 'C:/Windows/Fonts/arial.ttf'
                if os.path.exists(arial_path):
                    pdfmetrics.registerFont(TTFont(font_name, arial_path))
                    font_registered = True
            except:
                font_registered = False
            
            if not font_registered:
                font_name = 'Helvetica'
            
            # Minimal header - one line
            styles = getSampleStyleSheet()
            
            if exam_info:
                header_text = f"<b>{exam_info.get('ders_kodu', '')} - {exam_info.get('ders_adi', '')}</b> | {exam_info.get('tarih_saat', '')}"
                header_style = ParagraphStyle(
                    'CompactHeader',
                    parent=styles['Normal'],
                    fontSize=10,
                    fontName=font_name,
                    spaceAfter=4,
                    alignment=TA_CENTER
                )
                elements.append(Paragraph(header_text, header_style))
                elements.append(Spacer(1, 0.2*cm))
            
            # Group students by classroom
            students_by_classroom = defaultdict(list)
            for ogrenci_no, seat_info in seating_data.items():
                derslik_id = seat_info['derslik_id']
                students_by_classroom[derslik_id].append({
                    'ogrenci_no': ogrenci_no,
                    'ad_soyad': seat_info['ad_soyad'],
                    'sira': seat_info['sira'],
                    'sutun': seat_info['sutun']
                })
            
            # Create seating plan for each classroom
            for classroom in classrooms:
                derslik_id = classroom['derslik_id']
                derslik_adi = classroom.get('derslik_adi', classroom.get('derslik_kodu'))
                rows = classroom.get('satir_sayisi', 10)
                cols = classroom.get('sutun_sayisi', 6)
                sira_yapisi = classroom.get('sira_yapisi', 3)  # 2'li veya 3'lü grup
                
                students = students_by_classroom.get(derslik_id, [])
                if not students:
                    continue
                
                # Classroom title - inline
                classroom_title = Paragraph(
                    f"<b>{derslik_adi}</b> - {len(students)} öğrenci",
                    ParagraphStyle('ClassroomTitle', parent=styles['Normal'], 
                                   fontName=font_name, fontSize=10, spaceAfter=2)
                )
                elements.append(classroom_title)
                elements.append(Spacer(1, 0.1*cm))
                
                # Create seat grid
                seat_lookup = {}
                for student in students:
                    key = (student['sira'], student['sutun'])
                    seat_lookup[key] = student
                
                # Build table data for seating grid with CORRIDORS, window, and door
                grid_data = []
                
                # Calculate corridor positions based on sira_yapisi
                corridor_positions = []
                for c in range(1, cols):
                    if c % sira_yapisi == 0:
                        corridor_positions.append(c)
                
                # Simple header row with TAHTA spanning all
                num_visual_cols = cols + len(corridor_positions) + 2  # seats + corridors + window + door
                
                # Board row (spanning all columns)
                board_row = ['TAHTA →'] * num_visual_cols
                grid_data.append(board_row)
                
                # Only show used rows
                max_used_row = max([s['sira'] for s in students]) if students else rows
                display_rows = min(max_used_row + 1, rows)
                
                for row in range(1, display_rows + 1):
                    row_data = ['🪟']  # Window (first column)
                    
                    for col in range(1, cols + 1):
                        key = (row, col)
                        if key in seat_lookup:
                            student = seat_lookup[key]
                            # Longer names, better formatting
                            ad_soyad = student['ad_soyad']
                            # Split name smartly if too long
                            if len(ad_soyad) > 15:
                                parts = ad_soyad.split()
                                if len(parts) >= 2:
                                    cell_text = f"{student['ogrenci_no']}\n{parts[0][:8]}\n{parts[-1][:8]}"
                                else:
                                    cell_text = f"{student['ogrenci_no']}\n{ad_soyad[:15]}"
                            else:
                                cell_text = f"{student['ogrenci_no']}\n{ad_soyad}"
                            row_data.append(cell_text)
                        else:
                            row_data.append('--')  # Empty seat marker
                        
                        # Add corridor separator
                        if col in corridor_positions:
                            row_data.append('  ')  # Empty corridor space
                    
                    # Door (last column)
                    row_data.append('🚪' if row == display_rows // 2 else '')
                    grid_data.append(row_data)
                
                # Calculate for A4 landscape: better resolution
                available_width = 28 * cm
                available_height = 18 * cm
                
                # Calculate sizes
                num_corridors = len(corridor_positions)
                total_visual_cols = cols + num_corridors + 2
                
                # Larger cells for better readability
                cell_width = min(2.5*cm, (available_width - 1.6*cm - num_corridors*0.5*cm) / cols)
                cell_height = min(1.5*cm, (available_height - 1*cm) / (display_rows + 1))
                corridor_width = 0.5 * cm  # Wider corridor
                
                # Larger font for better resolution
                font_size = 8 if display_rows <= 8 else 7 if display_rows <= 12 else 6.5
                
                # Build column widths - optimized
                col_widths = [0.7*cm]  # Window (narrow)
                for c in range(1, cols + 1):
                    col_widths.append(cell_width)
                    if c in corridor_positions:
                        col_widths.append(corridor_width)
                col_widths.append(0.7*cm)  # Door (narrow)
                
                # Taller rows for better readability
                row_heights = [0.8*cm] + [cell_height] * display_rows
                
                grid_table = Table(grid_data, colWidths=col_widths, rowHeights=row_heights)
                
                # High-quality table style
                table_style = TableStyle([
                    # Board row (first row) - span entire width
                    ('SPAN', (0, 0), (-1, 0)),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fef3c7')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#92400e')),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('FONTNAME', (0, 0), (-1, 0), font_name),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    
                    # All cells - better spacing
                    ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                    ('FONTNAME', (0, 1), (-1, -1), font_name),
                    ('FONTSIZE', (0, 1), (-1, -1), font_size),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BOX', (0, 0), (-1, -1), 2, colors.black),
                    ('LEFTPADDING', (0, 0), (-1, -1), 2),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                    
                    # Window column (emoji)
                    ('FONTSIZE', (0, 1), (0, -1), 14),
                    
                    # Door column (emoji)
                    ('FONTSIZE', (-1, 1), (-1, -1), 14),
                ])
                
                # Color occupied seats (light blue) and empty seats (white)
                for row_idx in range(1, len(grid_data)):
                    for col_idx in range(1, len(grid_data[row_idx]) - 1):  # Skip window and door
                        cell_value = grid_data[row_idx][col_idx]
                        
                        # Skip corridor cells
                        if cell_value == '  ':
                            continue
                        
                        # Occupied seats
                        if cell_value and cell_value != '--' and '\n' in str(cell_value):
                            table_style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#dbeafe'))
                        # Empty seats
                        elif cell_value == '--':
                            table_style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#f3f4f6'))
                            table_style.add('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#9ca3af'))
                
                # Color corridor columns - distinct separator
                corridor_col_indices = []
                visual_col_idx = 1  # Start after window
                for c in range(1, cols + 1):
                    visual_col_idx += 1
                    if c in corridor_positions:
                        corridor_col_indices.append(visual_col_idx)
                        visual_col_idx += 1
                
                # Corridor styling - light gray background
                for corridor_col_idx in corridor_col_indices:
                    for row_idx in range(1, len(grid_data)):
                        table_style.add('BACKGROUND', (corridor_col_idx, row_idx), (corridor_col_idx, row_idx), colors.HexColor('#e5e7eb'))
                        table_style.add('LINEAFTER', (corridor_col_idx-1, row_idx), (corridor_col_idx-1, row_idx), 2, colors.HexColor('#9ca3af'))  # Thicker line before corridor
                
                grid_table.setStyle(table_style)
                elements.append(grid_table)
                
                # Add page break after each classroom (except last)
                if classroom != classrooms[-1]:
                    elements.append(PageBreak())
            
            # Build PDF
            doc.build(elements)
            logger.info(f"Seating plan PDF exported: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Seating plan PDF export error: {e}", exc_info=True)
            return False
    
    @staticmethod
    def _export_seating_list_pdf(data: Dict, file_path: str) -> bool:
        """Export seating list (table format) to PDF"""
        try:
            exam_info = data.get('exam_info', {})
            seating_data = data.get('seating_data', {})
            classrooms = data.get('classrooms', [])
            title = data.get('title', 'Öğrenci Oturma Listesi')
            
            if not seating_data:
                logger.warning("No seating data to export")
                return False
            
            # Use A4 with minimal margins for maximum space
            doc = SimpleDocTemplate(
                file_path, 
                pagesize=A4,
                topMargin=0.5*cm,
                bottomMargin=0.5*cm,
                leftMargin=0.5*cm,
                rightMargin=0.5*cm
            )
            elements = []
            
            # Register Turkish font
            font_registered = False
            font_name = 'TurkishFont'
            try:
                arial_path = 'C:/Windows/Fonts/arial.ttf'
                if os.path.exists(arial_path):
                    pdfmetrics.registerFont(TTFont(font_name, arial_path))
                    font_registered = True
            except:
                font_registered = False
            
            if not font_registered:
                font_name = 'Helvetica'
            
            # MINIMAL header - everything in one line
            styles = getSampleStyleSheet()
            
            # Single line header with all info
            if exam_info:
                header_text = f"<b>{exam_info.get('ders_kodu', '')} - {exam_info.get('ders_adi', '')}</b> | {exam_info.get('tarih_saat', '')} | Toplam: {len(seating_data)} öğrenci"
                header_style = ParagraphStyle(
                    'CompactHeader',
                    parent=styles['Normal'],
                    fontSize=9,
                    fontName=font_name,
                    spaceAfter=3,
                    alignment=TA_CENTER
                )
                elements.append(Paragraph(header_text, header_style))
                elements.append(Spacer(1, 0.1*cm))
            
            # Group students by classroom
            students_by_classroom = defaultdict(list)
            for ogrenci_no, seat_info in seating_data.items():
                derslik_id = seat_info['derslik_id']
                students_by_classroom[derslik_id].append({
                    'ogrenci_no': ogrenci_no,
                    'ad_soyad': seat_info['ad_soyad'],
                    'sira': seat_info['sira'],
                    'sutun': seat_info['sutun']
                })
            
            # Create list table for each classroom
            for classroom in classrooms:
                derslik_id = classroom['derslik_id']
                derslik_adi = classroom.get('derslik_adi', classroom.get('derslik_kodu'))
                
                students = students_by_classroom.get(derslik_id, [])
                if not students:
                    continue
                
                # Sort by row and column
                students.sort(key=lambda x: (x['sira'], x['sutun']))
                
                # Classroom header - inline with table, no extra space
                # (Will be included as table title)
                
                # Create table data - classroom name ABOVE table, not in header
                # Add classroom title as separate paragraph first
                classroom_header = Paragraph(
                    f"<b>{derslik_adi}</b> ({len(students)} öğrenci)",
                    ParagraphStyle('RoomHeader', parent=styles['Normal'], 
                                   fontName=font_name, fontSize=9, spaceAfter=2)
                )
                elements.append(classroom_header)
                
                # Now create simple table header
                table_data = [['#', 'Öğrenci No', 'Ad Soyad', 'Sıra-Süt']]
                
                for idx, student in enumerate(students, 1):
                    table_data.append([
                        str(idx),
                        str(student['ogrenci_no']),
                        student['ad_soyad'],
                        f"{student['sira']}-{student['sutun']}"
                    ])
                
                # Calculate optimal font to fit EVERYTHING on one page
                num_rows = len(table_data) - 1
                available_height = 28 * cm  # Maximum with minimal margins
                
                # ULTRA AGGRESSIVE - must fit on page
                if num_rows <= 20:
                    font_size = 8
                    row_padding = 2.5
                elif num_rows <= 35:
                    font_size = 7
                    row_padding = 2
                elif num_rows <= 50:
                    font_size = 6.5
                    row_padding = 1.5
                elif num_rows <= 70:
                    font_size = 6
                    row_padding = 1
                else:
                    # Emergency: super tiny
                    font_size = 5.5
                    row_padding = 0.5
                
                # Optimized column widths
                list_table = Table(table_data, colWidths=[1*cm, 2.8*cm, 10.5*cm, 1.7*cm])
                list_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # No column center
                    ('ALIGN', (3, 0), (4, -1), 'CENTER'),  # Sıra/Sütun center
                    ('FONTNAME', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), font_size),  # Dynamic font size
                    ('FONTSIZE', (0, 0), (-1, 0), font_size + 1),  # Header slightly larger
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BOX', (0, 0), (-1, -1), 1, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 1), (-1, -1), row_padding),  # Dynamic padding
                    ('BOTTOMPADDING', (0, 1), (-1, -1), row_padding),
                ]))
                elements.append(list_table)
                
                # Add page break after each classroom (except last)
                if classroom != classrooms[-1]:
                    elements.append(PageBreak())
            
            # Build PDF
            doc.build(elements)
            logger.info(f"Seating list PDF exported: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Seating list PDF export error: {e}", exc_info=True)
            return False
    
    @staticmethod
    def export_seating_plan(sinav_data: Dict, oturma_data: List[Dict], file_path: str) -> bool:
        """Export seating plan to PDF"""
        data = {
            'type': 'oturma_plani',
            'title': f"Oturma Planı - {sinav_data.get('ders_kodu', '')}",
            'data': oturma_data,
            'options': {'include_signatures': True}
        }

        return ExportUtils.export_to_pdf(data, file_path)
