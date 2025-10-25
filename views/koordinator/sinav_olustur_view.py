"""
Sınav Oluştur (Exam Create) View
Professional interface for creating exam schedules
"""

import logging
from datetime import datetime, timedelta
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QComboBox, QDateTimeEdit, QMessageBox,
    QGroupBox, QFormLayout, QSpinBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QProgressBar, QCheckBox,
    QLineEdit, QScrollArea, QFileDialog, QTabWidget
)
from PySide6.QtCore import Qt, QDateTime, QThread, Signal
from PySide6.QtGui import QFont

from models.database import db
from models.sinav_model import SinavModel
from models.ders_model import DersModel
from models.derslik_model import DerslikModel
from controllers.sinav_controller import SinavController
from algorithms.sinav_planlama import SinavPlanlama

logger = logging.getLogger(__name__)


class SinavPlanlamaThread(QThread):
    """Thread for exam planning algorithm"""
    progress = Signal(int, str)
    finished = Signal(dict)
    error = Signal(str)
    
    def __init__(self, params):
        super().__init__()
        self.params = params
    
    def run(self):
        try:
            planlama = SinavPlanlama()
            result = planlama.plan_exam_schedule(self.params, progress_callback=self.progress.emit)
            self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))


class SinavOlusturView(QWidget):
    """Exam schedule creation view"""
    
    def __init__(self, user_data, parent=None):
        super().__init__(parent)
        self.user_data = user_data
        self.bolum_id = user_data.get('bolum_id', 1)
        
        self.sinav_model = SinavModel(db)
        self.ders_model = DersModel(db)
        self.derslik_model = DerslikModel(db)
        self.sinav_controller = SinavController(self.sinav_model, self.ders_model, self.derslik_model)
        
        self.setup_ui()
        self.load_data()
    
    def setup_ui(self):
        """Setup UI"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(20)
        
        # Header
        header = QFrame()
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(0, 0, 0, 0)
        
        title = QLabel("Sınav Programı Yönetimi 📅")
        title.setFont(QFont("Segoe UI", 20, QFont.Bold))
        
        header_layout.addWidget(title)
        header_layout.addStretch()
        
        layout.addWidget(header)
        
        # Tab widget
        self.tab_widget = QTabWidget()
        
        # Tab 1: Existing Programs
        self.programs_tab = QWidget()
        self.setup_programs_tab()
        self.tab_widget.addTab(self.programs_tab, "📋 Mevcut Programlar")
        
        # Tab 2: Create New Program
        self.create_tab = QWidget()
        self.setup_create_tab()
        self.tab_widget.addTab(self.create_tab, "➕ Yeni Program Oluştur")
        
        layout.addWidget(self.tab_widget)
    
    def setup_programs_tab(self):
        """Setup existing programs tab"""
        layout = QVBoxLayout(self.programs_tab)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)
        
        # Info
        info = QLabel("Oluşturulmuş sınav programlarını görüntüleyin ve yönetin:")
        info.setStyleSheet("color: #6b7280; font-size: 13px;")
        layout.addWidget(info)
        
        # Refresh button
        refresh_btn = QPushButton("🔄 Yenile")
        refresh_btn.setFixedHeight(36)
        refresh_btn.clicked.connect(self.load_existing_programs)
        layout.addWidget(refresh_btn, alignment=Qt.AlignRight)
        
        # Programs table
        self.programs_table = QTableWidget()
        self.programs_table.setColumnCount(6)
        self.programs_table.setHorizontalHeaderLabels([
            "Program Adı", "Sınav Tipi", "Başlangıç", "Bitiş", "Sınav Sayısı", "İşlemler"
        ])
        self.programs_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.programs_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.programs_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.programs_table.setAlternatingRowColors(True)
        header = self.programs_table.horizontalHeader()
        header.setSectionResizeMode(5, QHeaderView.Fixed)
        self.programs_table.setColumnWidth(5, 260)
        layout.addWidget(self.programs_table)
        
        # Load programs
        self.load_existing_programs()
    
    def setup_create_tab(self):
        """Setup create new program tab"""
        # Main scroll area for entire form
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("""
            QScrollArea {
                background: transparent;
                border: none;
            }
            QScrollBar:vertical {
                border: none;
                background: #f3f4f6;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background: #9ca3af;
                border-radius: 6px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover {
                background: #6b7280;
            }
        """)
        
        scroll_content = QWidget()
        layout = QVBoxLayout(scroll_content)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)
        
        # Create tab layout
        tab_layout = QVBoxLayout(self.create_tab)
        tab_layout.setContentsMargins(0, 0, 0, 0)
        tab_layout.addWidget(scroll)
        
        # Parameters - Compact 2-column layout
        params_card = QGroupBox("⚙️ Sınav Parametreleri")
        params_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                border: 2px solid #e5e7eb;
                border-radius: 10px;
                margin-top: 12px;
                padding: 16px;
                background: white;
            }
            QGroupBox::title {
                color: #1f2937;
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 8px;
            }
        """)
        params_layout = QHBoxLayout(params_card)
        params_layout.setSpacing(20)
        
        # Left column
        left_col = QFormLayout()
        left_col.setSpacing(12)
        left_col.setLabelAlignment(Qt.AlignRight)
        
        self.sinav_tipi_combo = QComboBox()
        self.sinav_tipi_combo.addItems(["Vize", "Final", "Bütünleme"])
        self.sinav_tipi_combo.setFixedHeight(36)
        left_col.addRow("📝 Sınav Tipi:", self.sinav_tipi_combo)
        
        self.baslangic_tarih = QDateTimeEdit()
        self.baslangic_tarih.setDateTime(QDateTime.currentDateTime().addDays(7))
        self.baslangic_tarih.setCalendarPopup(True)
        self.baslangic_tarih.setDisplayFormat("dd.MM.yyyy")
        self.baslangic_tarih.setFixedHeight(36)
        left_col.addRow("📅 Başlangıç:", self.baslangic_tarih)
        
        self.bitis_tarih = QDateTimeEdit()
        self.bitis_tarih.setDateTime(QDateTime.currentDateTime().addDays(14))
        self.bitis_tarih.setCalendarPopup(True)
        self.bitis_tarih.setDisplayFormat("dd.MM.yyyy")
        self.bitis_tarih.setFixedHeight(36)
        left_col.addRow("📅 Bitiş:", self.bitis_tarih)
        
        params_layout.addLayout(left_col, 1)
        
        # Right column
        right_col = QFormLayout()
        right_col.setSpacing(12)
        right_col.setLabelAlignment(Qt.AlignRight)
        
        self.sinav_suresi = QSpinBox()
        self.sinav_suresi.setMinimum(30)
        self.sinav_suresi.setMaximum(240)
        self.sinav_suresi.setValue(75)
        self.sinav_suresi.setSuffix(" dk")
        self.sinav_suresi.setFixedHeight(36)
        # Track and propagate changes to course-specific durations
        self._last_global_duration = self.sinav_suresi.value()
        self.sinav_suresi.valueChanged.connect(self.on_global_duration_changed)
        right_col.addRow("⏱️ Sınav Süresi:", self.sinav_suresi)
        
        self.ara_suresi = QSpinBox()
        self.ara_suresi.setMinimum(0)
        self.ara_suresi.setMaximum(120)
        self.ara_suresi.setValue(15)
        self.ara_suresi.setSuffix(" dk")
        self.ara_suresi.setFixedHeight(36)
        right_col.addRow("⏸️ Ara Süresi:", self.ara_suresi)

        # Per-class per-day exam limit (0 = limitsiz)
        self.class_per_day_limit = QSpinBox()
        self.class_per_day_limit.setMinimum(0)
        self.class_per_day_limit.setMaximum(10)
        self.class_per_day_limit.setValue(0)
        self.class_per_day_limit.setSuffix(" /gün")
        self.class_per_day_limit.setFixedHeight(36)
        right_col.addRow("🏷️ Sınıf/Gün Sınırı:", self.class_per_day_limit)

        # (Removed) Minimum rest minutes control – ara süresi yeterli

        # Minimum shared students to consider conflict
        self.min_conflict_overlap = QSpinBox()
        self.min_conflict_overlap.setMinimum(1)
        self.min_conflict_overlap.setMaximum(50)
        self.min_conflict_overlap.setValue(1)
        self.min_conflict_overlap.setFixedHeight(36)
        right_col.addRow("👥 Çakışma Eşiği:", self.min_conflict_overlap)

        # No parallel exams option
        from PySide6.QtWidgets import QCheckBox
        self.no_parallel_checkbox = QCheckBox("Sınavlar aynı anda olmasın")
        self.no_parallel_checkbox.setChecked(False)
        right_col.addRow("", self.no_parallel_checkbox)
        
        params_layout.addLayout(right_col, 1)
        
        layout.addWidget(params_card)
        
        # Weekday selection - Compact
        gunler_card = QGroupBox("📅 Sınav Günleri")
        gunler_card.setStyleSheet("""
            QGroupBox {
                font-size: 13px;
                font-weight: bold;
                border: 2px solid #e5e7eb;
                border-radius: 8px;
                margin-top: 10px;
                padding: 12px;
                background: #f9fafb;
            }
        """)
        gunler_layout = QHBoxLayout(gunler_card)
        gunler_layout.setSpacing(8)
        
        self.gun_checkboxes = {}
        gun_isimleri = {
            0: "Pzt", 1: "Sal", 2: "Çar", 3: "Per", 4: "Cum", 5: "Cmt", 6: "Paz"
        }
        
        for day_num, day_name in gun_isimleri.items():
            checkbox = QCheckBox(day_name)
            checkbox.setStyleSheet("""
                QCheckBox {
                    font-size: 12px;
                    font-weight: bold;
                    spacing: 5px;
                }
                QCheckBox::indicator {
                    width: 18px;
                    height: 18px;
                }
            """)
            if day_num < 5:
                checkbox.setChecked(True)
            self.gun_checkboxes[day_num] = checkbox
            gunler_layout.addWidget(checkbox)
        
        gunler_layout.addStretch()
        layout.addWidget(gunler_card)
        
        # Time constraints - Compact 2x2 grid
        zaman_card = QGroupBox("🕐 Sınav Saatleri")
        zaman_card.setStyleSheet("""
            QGroupBox {
                font-size: 13px;
                font-weight: bold;
                border: 2px solid #e5e7eb;
                border-radius: 8px;
                margin-top: 10px;
                padding: 12px;
                background: white;
            }
        """)
        zaman_layout = QFormLayout(zaman_card)
        zaman_layout.setSpacing(10)
        zaman_layout.setLabelAlignment(Qt.AlignRight)
        
        # First exam time
        ilk_sinav_layout = QHBoxLayout()
        ilk_sinav_layout.setSpacing(4)
        self.ilk_sinav_saat = QSpinBox()
        self.ilk_sinav_saat.setMinimum(0)
        self.ilk_sinav_saat.setMaximum(23)
        self.ilk_sinav_saat.setValue(10)
        self.ilk_sinav_saat.setFixedWidth(60)
        self.ilk_sinav_dakika = QSpinBox()
        self.ilk_sinav_dakika.setMinimum(0)
        self.ilk_sinav_dakika.setMaximum(59)
        self.ilk_sinav_dakika.setValue(0)
        self.ilk_sinav_dakika.setFixedWidth(60)
        ilk_sinav_layout.addWidget(self.ilk_sinav_saat)
        ilk_sinav_layout.addWidget(QLabel(":"))
        ilk_sinav_layout.addWidget(self.ilk_sinav_dakika)
        ilk_sinav_layout.addStretch()
        zaman_layout.addRow("⏰ İlk Sınav:", ilk_sinav_layout)
        
        # Last exam time
        son_sinav_layout = QHBoxLayout()
        son_sinav_layout.setSpacing(4)
        self.son_sinav_saat = QSpinBox()
        self.son_sinav_saat.setMinimum(0)
        self.son_sinav_saat.setMaximum(23)
        self.son_sinav_saat.setValue(19)
        self.son_sinav_saat.setFixedWidth(60)
        self.son_sinav_dakika = QSpinBox()
        self.son_sinav_dakika.setMinimum(0)
        self.son_sinav_dakika.setMaximum(59)
        self.son_sinav_dakika.setValue(15)
        self.son_sinav_dakika.setFixedWidth(60)
        son_sinav_layout.addWidget(self.son_sinav_saat)
        son_sinav_layout.addWidget(QLabel(":"))
        son_sinav_layout.addWidget(self.son_sinav_dakika)
        son_sinav_layout.addStretch()
        zaman_layout.addRow("🔚 Son Sınav:", son_sinav_layout)
        
        # Lunch break
        ogle_bas_layout = QHBoxLayout()
        ogle_bas_layout.setSpacing(4)
        self.ogle_baslangic_saat = QSpinBox()
        self.ogle_baslangic_saat.setMinimum(0)
        self.ogle_baslangic_saat.setMaximum(23)
        self.ogle_baslangic_saat.setValue(12)
        self.ogle_baslangic_saat.setFixedWidth(60)
        self.ogle_baslangic_dakika = QSpinBox()
        self.ogle_baslangic_dakika.setMinimum(0)
        self.ogle_baslangic_dakika.setMaximum(59)
        self.ogle_baslangic_dakika.setValue(0)
        self.ogle_baslangic_dakika.setFixedWidth(60)
        ogle_bas_layout.addWidget(self.ogle_baslangic_saat)
        ogle_bas_layout.addWidget(QLabel(":"))
        ogle_bas_layout.addWidget(self.ogle_baslangic_dakika)
        ogle_bas_layout.addStretch()
        zaman_layout.addRow("🍽️ Öğle Arası Baş:", ogle_bas_layout)
        
        ogle_bit_layout = QHBoxLayout()
        ogle_bit_layout.setSpacing(4)
        self.ogle_bitis_saat = QSpinBox()
        self.ogle_bitis_saat.setMinimum(0)
        self.ogle_bitis_saat.setMaximum(23)
        self.ogle_bitis_saat.setValue(13)
        self.ogle_bitis_saat.setFixedWidth(60)
        self.ogle_bitis_dakika = QSpinBox()
        self.ogle_bitis_dakika.setMinimum(0)
        self.ogle_bitis_dakika.setMaximum(59)
        self.ogle_bitis_dakika.setValue(30)
        self.ogle_bitis_dakika.setFixedWidth(60)
        ogle_bit_layout.addWidget(self.ogle_bitis_saat)
        ogle_bit_layout.addWidget(QLabel(":"))
        ogle_bit_layout.addWidget(self.ogle_bitis_dakika)
        ogle_bit_layout.addStretch()
        zaman_layout.addRow("🍽️ Öğle Arası Bit:", ogle_bit_layout)
        
        layout.addWidget(zaman_card)
        
        # Course selection
        ders_card = QGroupBox("📚 Sınavı Yapılacak Dersler")
        ders_card.setStyleSheet("""
            QGroupBox {
                font-size: 13px;
                font-weight: bold;
                border: 2px solid #e5e7eb;
                border-radius: 8px;
                margin-top: 10px;
                padding: 12px;
                background: white;
            }
        """)
        ders_layout = QVBoxLayout(ders_card)
        
        # Search and select all
        ders_toolbar = QHBoxLayout()
        ders_toolbar.setSpacing(8)
        
        self.ders_search = QLineEdit()
        self.ders_search.setPlaceholderText("🔍 Ders ara...")
        self.ders_search.setFixedHeight(36)
        self.ders_search.textChanged.connect(self.filter_courses)
        
        select_all_btn = QPushButton("☑️ Tümünü Seç/Kaldır")
        select_all_btn.setFixedHeight(36)
        select_all_btn.setFixedWidth(160)
        select_all_btn.clicked.connect(self.toggle_all_courses)
        
        ders_toolbar.addWidget(self.ders_search)
        ders_toolbar.addWidget(select_all_btn)
        ders_layout.addLayout(ders_toolbar)
        
        # Scrollable course list with duration controls
        ders_scroll = QScrollArea()
        ders_scroll.setWidgetResizable(True)
        ders_scroll.setMinimumHeight(400)
        ders_scroll.setStyleSheet("""
            QScrollArea {
                border: 1px solid #e5e7eb;
                border-radius: 6px;
                background: #f9fafb;
            }
        """)
        
        self.ders_container = QWidget()
        self.ders_container_layout = QVBoxLayout(self.ders_container)
        self.ders_container_layout.setSpacing(4)
        self.ders_container_layout.setContentsMargins(8, 8, 8, 8)
        self.ders_checkboxes = {}
        self.ders_duration_spinboxes = {}  # Store duration spinboxes
        
        ders_scroll.setWidget(self.ders_container)
        ders_layout.addWidget(ders_scroll)
        
        # Stats
        self.ders_stats_label = QLabel("📊 Yükleniyor...")
        self.ders_stats_label.setStyleSheet("""
            color: #6b7280;
            font-size: 12px;
            font-weight: bold;
            padding: 8px;
            background: #f9fafb;
            border-radius: 6px;
        """)
        ders_layout.addWidget(self.ders_stats_label)
        
        layout.addWidget(ders_card)
        
        # Progress
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFixedHeight(32)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #e5e7eb;
                border-radius: 8px;
                text-align: center;
                font-weight: bold;
                background: white;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3b82f6, stop:1 #2563eb);
                border-radius: 6px;
            }
        """)
        layout.addWidget(self.progress_bar)
        
        self.progress_label = QLabel()
        self.progress_label.setVisible(False)
        self.progress_label.setStyleSheet("color: #6b7280; font-size: 13px; font-weight: bold;")
        self.progress_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.progress_label)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        self.create_btn = QPushButton("🚀 Sınav Programı Oluştur")
        self.create_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3b82f6, stop:1 #2563eb);
                color: white;
                border: none;
                border-radius: 10px;
                font-size: 15px;
                font-weight: bold;
                padding: 14px 32px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #2563eb, stop:1 #1e40af);
            }
            QPushButton:pressed {
                background: #1e40af;
            }
        """)
        self.create_btn.setFixedHeight(50)
        self.create_btn.setFixedWidth(260)
        self.create_btn.setCursor(Qt.PointingHandCursor)
        self.create_btn.clicked.connect(self.create_schedule)
        
        button_layout.addStretch()
        button_layout.addWidget(self.create_btn)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        layout.addSpacing(20)
        
        # Set scroll widget
        scroll.setWidget(scroll_content)
        
        # Results section (not in scroll, added to main tab layout)
        self.results_group = QGroupBox("📊 Oluşturulan Program")
        self.results_group.setVisible(False)
        self.results_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                border: 2px solid #10b981;
                border-radius: 10px;
                margin-top: 12px;
                padding: 16px;
                background: #f0fdf4;
            }
        """)
        results_layout = QVBoxLayout(self.results_group)
        
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(6)
        self.results_table.setHorizontalHeaderLabels([
            "Tarih/Saat", "Ders Kodu", "Ders Adı", "Öğretim Elemanı", "Derslik", "Öğrenci Sayısı"
        ])
        self.results_table.verticalHeader().setVisible(False)
        self.results_table.setAlternatingRowColors(True)
        
        header = self.results_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)
        header.setSectionResizeMode(1, QHeaderView.Fixed)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(3, QHeaderView.Fixed)
        header.setSectionResizeMode(4, QHeaderView.Fixed)
        header.setSectionResizeMode(5, QHeaderView.Fixed)
        
        self.results_table.setColumnWidth(0, 150)
        self.results_table.setColumnWidth(1, 100)
        self.results_table.setColumnWidth(3, 200)
        self.results_table.setColumnWidth(4, 100)
        self.results_table.setColumnWidth(5, 120)
        
        results_layout.addWidget(self.results_table)
        
        btn_layout = QHBoxLayout()
        
        save_btn = QPushButton("💾 Programı Kaydet")
        save_btn.setObjectName("primaryBtn")
        save_btn.setFixedHeight(40)
        save_btn.clicked.connect(self.save_schedule)
        
        export_btn = QPushButton("📊 Excel'e Aktar")
        export_btn.setObjectName("secondaryBtn")
        export_btn.setFixedHeight(40)
        export_btn.clicked.connect(self.export_to_excel)
        
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(export_btn)
        
        results_layout.addLayout(btn_layout)
        
        # Add results to main tab (not inside scroll)
        tab_layout.addWidget(self.results_group)
    
    def load_existing_programs(self):
        """Load and display existing programs"""
        try:
            programs = self.sinav_model.get_programs_by_bolum(self.bolum_id)
            
            self.programs_table.setRowCount(0)
            
            for program in programs:
                row = self.programs_table.rowCount()
                self.programs_table.insertRow(row)
                
                # Get exam count
                sinavlar = self.sinav_model.get_sinavlar_by_program(program['program_id'])
                exam_count = len(sinavlar)
                
                self.programs_table.setItem(row, 0, QTableWidgetItem(program['program_adi']))
                self.programs_table.setItem(row, 1, QTableWidgetItem(program.get('sinav_tipi', 'Final')))
                self.programs_table.setItem(row, 2, QTableWidgetItem(str(program.get('baslangic_tarihi', ''))))
                self.programs_table.setItem(row, 3, QTableWidgetItem(str(program.get('bitis_tarihi', ''))))
                self.programs_table.setItem(row, 4, QTableWidgetItem(str(exam_count)))
                
                # Action buttons
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(8, 4, 8, 4)
                actions_layout.setSpacing(8)
                
                view_btn = QPushButton("📋 Görüntüle")
                view_btn.setFixedHeight(32)
                view_btn.setMinimumWidth(110)
                view_btn.clicked.connect(lambda checked=False, p=dict(program): self.view_program(p))
                
                delete_btn = QPushButton("🗑️ Sil")
                delete_btn.setFixedHeight(32)
                delete_btn.setMinimumWidth(90)
                delete_btn.setStyleSheet("background-color: #ef4444; color: white;")
                delete_btn.clicked.connect(lambda checked=False, p=dict(program): self.delete_program(p))
                
                actions_layout.addWidget(view_btn)
                actions_layout.addWidget(delete_btn)
                actions_layout.addStretch()
                
                self.programs_table.setCellWidget(row, 5, actions_widget)
            
            logger.info(f"Loaded {len(programs)} exam programs")
            
        except Exception as e:
            logger.error(f"Error loading programs: {e}", exc_info=True)
            QMessageBox.critical(self, "Hata", f"Programlar yüklenirken hata:\n{str(e)}")
    
    def view_program(self, program):
        """View program details"""
        try:
            sinavlar = self.sinav_model.get_sinavlar_by_program(program['program_id'])
            
            if not sinavlar:
                QMessageBox.information(self, "Bilgi", "Bu programda henüz sınav yok!")
                return
            
            # Create message with exam details
            message = f"📅 {program['program_adi']}\n"
            message += f"📝 Tip: {program.get('sinav_tipi', 'Final')}\n"
            message += f"📆 {program.get('baslangic_tarihi')} - {program.get('bitis_tarihi')}\n\n"
            message += f"Toplam {len(sinavlar)} sınav:\n\n"
            
            for sinav in sinavlar[:10]:  # Show first 10
                message += f"• {sinav['ders_kodu']}: {sinav['tarih_saat']} - {sinav.get('derslik_kodu', 'N/A')}\n"
            
            if len(sinavlar) > 10:
                message += f"\n... ve {len(sinavlar) - 10} sınav daha"
            
            QMessageBox.information(self, "Program Detayları", message)
            
        except Exception as e:
            logger.error(f"Error viewing program: {e}", exc_info=True)
            QMessageBox.critical(self, "Hata", f"Program görüntülenirken hata:\n{str(e)}")
    
    def delete_program(self, program):
        """Delete a program"""
        reply = QMessageBox.question(
            self,
            "Programı Sil",
            f"'{program['program_adi']}' programını silmek istediğinizden emin misiniz?\n\n"
            f"Bu işlem geri alınamaz ve programa ait tüm sınavlar silinecektir!",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                result = self.sinav_model.delete_program(program['program_id'])
                
                if result:
                    QMessageBox.information(self, "Başarılı", "Program başarıyla silindi!")
                    self.load_existing_programs()
                else:
                    QMessageBox.warning(self, "Uyarı", "Program silinemedi!")
                    
            except Exception as e:
                logger.error(f"Error deleting program: {e}", exc_info=True)
                QMessageBox.critical(self, "Hata", f"Program silinirken hata:\n{str(e)}")
    
    def load_data(self):
        """Load necessary data"""
        try:
            # Load courses and classrooms to verify availability
            dersler = self.ders_model.get_dersler_by_bolum(self.bolum_id)
            derslikler = self.derslik_model.get_derslikler_by_bolum(self.bolum_id)
            
            if not dersler:
                QMessageBox.warning(self, "Uyarı", "Henüz ders tanımlanmamış!")
                return
            
            if not derslikler:
                QMessageBox.warning(self, "Uyarı", "Henüz derslik tanımlanmamış!")
                return
            
            # Populate course checkboxes
            self.populate_course_list(dersler)
                
        except Exception as e:
            logger.error(f"Error loading data: {e}")
    
    def populate_course_list(self, dersler):
        """Populate course selection checkboxes with duration controls"""
        # Clear existing
        while self.ders_container_layout.count():
            item = self.ders_container_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        self.ders_checkboxes.clear()
        self.ders_duration_spinboxes.clear()
        
        for ders in dersler:
            # Create horizontal layout for each course
            course_row = QWidget()
            course_layout = QHBoxLayout(course_row)
            course_layout.setContentsMargins(0, 0, 0, 0)
            course_layout.setSpacing(10)
            
            # Checkbox for course selection
            checkbox = QCheckBox(f"{ders['ders_kodu']} - {ders['ders_adi']}")
            checkbox.setChecked(True)  # All selected by default
            checkbox.setProperty('ders_id', ders['ders_id'])
            checkbox.setProperty('ders_kodu', ders['ders_kodu'])
            checkbox.stateChanged.connect(self.update_course_stats)
            checkbox.setMinimumWidth(400)
            
            # Duration spinbox for this course
            duration_label = QLabel("Süre:")
            duration_spinbox = QSpinBox()
            duration_spinbox.setMinimum(30)
            duration_spinbox.setMaximum(240)
            # Initialize from current global duration
            duration_spinbox.setValue(self.sinav_suresi.value())
            duration_spinbox.setSuffix(" dk")
            duration_spinbox.setFixedWidth(90)
            
            course_layout.addWidget(checkbox)
            course_layout.addStretch()
            course_layout.addWidget(duration_label)
            course_layout.addWidget(duration_spinbox)
            
            self.ders_checkboxes[ders['ders_id']] = checkbox
            self.ders_duration_spinboxes[ders['ders_id']] = duration_spinbox
            self.ders_container_layout.addWidget(course_row)
        
        self.update_course_stats()
    
    def filter_courses(self):
        """Filter courses based on search"""
        search_text = self.ders_search.text().lower()
        
        for ders_id, checkbox in self.ders_checkboxes.items():
            text = checkbox.text().lower()
            # Show/hide the parent widget (course_row) instead of just checkbox
            parent_widget = checkbox.parent()
            if parent_widget:
                parent_widget.setVisible(search_text in text)
    
    def toggle_all_courses(self):
        """Toggle all course selections"""
        # Check if all are selected
        all_checked = all(cb.isChecked() for cb in self.ders_checkboxes.values())
        
        # Toggle
        for checkbox in self.ders_checkboxes.values():
            checkbox.setChecked(not all_checked)
        
        self.update_course_stats()
    
    def update_course_stats(self):
        """Update course selection statistics"""
        total = len(self.ders_checkboxes)
        selected = sum(1 for cb in self.ders_checkboxes.values() if cb.isChecked())
        self.ders_stats_label.setText(f"Seçili: {selected} / {total} ders")
    
    def create_schedule(self):
        """Create exam schedule"""
        # Validate dates
        if self.baslangic_tarih.dateTime() >= self.bitis_tarih.dateTime():
            QMessageBox.warning(self, "Uyarı", "Bitiş tarihi başlangıç tarihinden sonra olmalıdır!")
            return
        
        # Get allowed weekdays
        allowed_weekdays = [day for day, checkbox in self.gun_checkboxes.items() if checkbox.isChecked()]
        
        if not allowed_weekdays:
            QMessageBox.warning(self, "Uyarı", "En az bir gün seçmelisiniz!")
            return
        
        # Get selected courses
        selected_ders_ids = [ders_id for ders_id, checkbox in self.ders_checkboxes.items() if checkbox.isChecked()]
        
        if not selected_ders_ids:
            QMessageBox.warning(self, "Uyarı", "En az bir ders seçmelisiniz!")
            return
        
        # Collect custom exam durations for each course
        ders_sinavlari_suresi = {}
        for ders_id in selected_ders_ids:
            if ders_id in self.ders_duration_spinboxes:
                course_duration = self.ders_duration_spinboxes[ders_id].value()
                # Only include custom duration if it differs from global, so global applies by default
                if course_duration != self.sinav_suresi.value():
                    ders_sinavlari_suresi[ders_id] = course_duration
        
        # Format time strings
        ilk_sinav = f"{self.ilk_sinav_saat.value():02d}:{self.ilk_sinav_dakika.value():02d}"
        son_sinav = f"{self.son_sinav_saat.value():02d}:{self.son_sinav_dakika.value():02d}"
        ogle_baslangic = f"{self.ogle_baslangic_saat.value():02d}:{self.ogle_baslangic_dakika.value():02d}"
        ogle_bitis = f"{self.ogle_bitis_saat.value():02d}:{self.ogle_bitis_dakika.value():02d}"
        
        params = {
            'bolum_id': self.bolum_id,
            'sinav_tipi': self.sinav_tipi_combo.currentText(),
            'baslangic_tarih': self.baslangic_tarih.dateTime().toPython(),
            'bitis_tarih': self.bitis_tarih.dateTime().toPython(),
            'varsayilan_sinav_suresi': self.sinav_suresi.value(),
            'ara_suresi': self.ara_suresi.value(),
            'allowed_weekdays': allowed_weekdays,
            'selected_ders_ids': selected_ders_ids,
            'ders_sinavlari_suresi': ders_sinavlari_suresi,
            'gunluk_ilk_sinav': ilk_sinav,
            'gunluk_son_sinav': son_sinav,
            'ogle_arasi_baslangic': ogle_baslangic,
            'ogle_arasi_bitis': ogle_bitis,
            'class_per_day_limit': self.class_per_day_limit.value(),
            'min_conflict_overlap': self.min_conflict_overlap.value(),
            'no_parallel_exams': self.no_parallel_checkbox.isChecked(),
        }
        
        # Show progress
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.progress_label.setVisible(True)
        self.progress_label.setText("Sınav programı oluşturuluyor...")
        self.create_btn.setEnabled(False)
        
        # Start planning thread
        self.planning_thread = SinavPlanlamaThread(params)
        self.planning_thread.progress.connect(self.on_planning_progress)
        self.planning_thread.finished.connect(self.on_planning_finished)
        self.planning_thread.error.connect(self.on_planning_error)
        self.planning_thread.start()

    def on_global_duration_changed(self, new_value: int):
        """When the global exam duration changes, update per-course durations that
        were previously equal to the old global value (preserve explicit overrides)."""
        try:
            # If course widgets not yet built, nothing to update
            if not hasattr(self, 'ders_duration_spinboxes'):
                self._last_global_duration = new_value
                return
            for ders_id, sb in self.ders_duration_spinboxes.items():
                if sb.value() == self._last_global_duration:
                    sb.setValue(new_value)
            self._last_global_duration = new_value
        except Exception as e:
            logger.error(f"Error syncing global duration: {e}")
    
    def on_planning_progress(self, percent, message):
        """Update planning progress"""
        self.progress_bar.setValue(percent)
        self.progress_label.setText(message)
    
    def on_planning_finished(self, result):
        """Handle planning completion"""
        self.progress_bar.setVisible(False)
        self.progress_label.setVisible(False)
        self.create_btn.setEnabled(True)
        
        schedule = result.get('schedule', [])
        
        if schedule:
            # Show partial or complete schedule
            self.current_schedule = schedule
            self.display_schedule(schedule)
        
        # Count unique exams (ders + datetime) to avoid overcounting per-classroom rows
        unique_exam_keys = set()
        for s in schedule:
            ts = s['tarih_saat']
            key_dt = ts if not isinstance(ts, str) else datetime.fromisoformat(ts)
            unique_exam_keys.add((s.get('ders_id'), key_dt))
        unique_exam_count = len(unique_exam_keys)

        if result.get('success'):
            QMessageBox.information(
                self,
                "Başarılı",
                f"✅ Sınav programı başarıyla oluşturuldu!\n\n"
                f"Toplam {unique_exam_count} sınav planlandı."
            )
        else:
            # Show warning with details and partial schedule
            unassigned = result.get('unassigned_courses', [])
            message = result.get('message', 'Program oluşturulamadı!')
            
            if schedule:
                message += f"\n\n✅ {unique_exam_count} sınav yerleştirildi."
                message += f"\n❌ {len(unassigned)} ders yerleştirilemedi."
            
            QMessageBox.warning(self, "Kısmi Program Oluşturuldu", message)
    
    def on_planning_error(self, error_msg):
        """Handle planning error"""
        self.progress_bar.setVisible(False)
        self.progress_label.setVisible(False)
        self.create_btn.setEnabled(True)
        
        QMessageBox.critical(self, "Hata", f"Program oluşturulurken hata oluştu:\n{error_msg}")
    
    def display_schedule(self, schedule):
        """Display created schedule"""
        self.results_group.setVisible(True)
        self.results_table.setRowCount(0)
        
        for row, sinav in enumerate(schedule):
            self.results_table.insertRow(row)
            
            tarih = datetime.fromisoformat(sinav['tarih_saat']) if isinstance(sinav['tarih_saat'], str) else sinav['tarih_saat']
            tarih_str = tarih.strftime("%d.%m.%Y %H:%M")
            
            self.results_table.setItem(row, 0, QTableWidgetItem(tarih_str))
            self.results_table.setItem(row, 1, QTableWidgetItem(sinav.get('ders_kodu', '')))
            self.results_table.setItem(row, 2, QTableWidgetItem(sinav.get('ders_adi', '')))
            self.results_table.setItem(row, 3, QTableWidgetItem(sinav.get('ogretim_elemani', '')))
            # Show derslik_adi if available, otherwise derslik_kodu
            derslik_display = sinav.get('derslik_adi', sinav.get('derslik_kodu', ''))
            self.results_table.setItem(row, 4, QTableWidgetItem(derslik_display))
            
            ogrenci_item = QTableWidgetItem(str(sinav.get('ogrenci_sayisi', 0)))
            ogrenci_item.setTextAlignment(Qt.AlignCenter)
            self.results_table.setItem(row, 5, ogrenci_item)
    
    def save_schedule(self):
        """Save schedule to database"""
        if not hasattr(self, 'current_schedule') or not self.current_schedule:
            QMessageBox.warning(self, "Uyarı", "Kaydedilecek program bulunamadı!")
            return
        
        reply = QMessageBox.question(
            self,
            "Programı Kaydet",
            f"{len(self.current_schedule)} sınavı veritabanına kaydetmek istediğinizden emin misiniz?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        
        if reply == QMessageBox.Yes:
            try:
                result = self.sinav_controller.save_exam_schedule(self.current_schedule)
                
                if result['success']:
                    QMessageBox.information(self, "Başarılı", result['message'])
                    self.current_schedule = []
                    self.results_group.setVisible(False)
                else:
                    QMessageBox.warning(self, "Hata", result['message'])
                    
            except Exception as e:
                logger.error(f"Error saving schedule: {e}")
                QMessageBox.critical(self, "Hata", f"Program kaydedilirken hata oluştu:\n{str(e)}")
    
    def export_to_excel(self):
        """Export schedule to Excel file with professional formatting (like the example)"""
        if not hasattr(self, 'current_schedule') or not self.current_schedule:
            QMessageBox.warning(self, "Uyarı", "Aktarılacak program bulunamadı!")
            return
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from collections import defaultdict
        
        # Ask for save location
        default_name = f"sinav_programi_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Excel Dosyası Kaydet",
            default_name,
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            # Group schedule by date and time
            schedule_by_datetime = defaultdict(list)
            for sinav in self.current_schedule:
                tarih = datetime.fromisoformat(sinav['tarih_saat']) if isinstance(sinav['tarih_saat'], str) else sinav['tarih_saat']
                datetime_key = (tarih.date(), tarih.time())
                schedule_by_datetime[datetime_key].append(sinav)
            
            # Sort by datetime
            sorted_datetimes = sorted(schedule_by_datetime.keys())
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Sınav Programı"
            
            # Get bolum name for title
            bolum_adi = "BİLGİSAYAR MÜHENDİSLİĞİ BÖLÜMÜ"  # You can make this dynamic
            sinav_tipi = self.current_schedule[0].get('sinav_tipi', 'SINAV')
            
            # Title row - merged cells with orange background
            ws.merge_cells('A1:E1')
            title_cell = ws['A1']
            title_cell.value = f"{bolum_adi} {sinav_tipi.upper()} SINAV PROGRAMI"
            title_cell.font = Font(size=14, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 25
            
            # Header row
            headers = ['Tarih', 'Sınav Saati', 'Ders Adı', 'Öğretim Elemanı', 'Derslik']
            ws.append(headers)
            
            # Style header row
            header_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in ws[2]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            ws.row_dimensions[2].height = 20
            
            # Border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Data rows
            current_row = 3
            date_merge_start = {}  # Track where each date starts for merging
            
            for datetime_key in sorted_datetimes:
                date_obj, time_obj = datetime_key
                exams_at_this_time = schedule_by_datetime[datetime_key]
                
                # Group exams by course (multiple classrooms for same course)
                course_groups = defaultdict(list)
                for exam in exams_at_this_time:
                    course_key = (exam['ders_id'], exam['ders_kodu'], exam['ders_adi'])
                    course_groups[course_key].append(exam)
                
                date_str = date_obj.strftime('%d.%m.%Y')
                time_str = time_obj.strftime('%H.%M')
                
                # Track start row for this date
                if date_str not in date_merge_start:
                    date_merge_start[date_str] = current_row
                
                # Add each course
                for (ders_id, ders_kodu, ders_adi), course_exams in course_groups.items():
                    # Combine all classroom NAMES (not codes) for this course
                    derslikler = '-'.join([exam.get('derslik_adi', exam.get('derslik_kodu', '')) for exam in course_exams])
                    
                    # Get instructor from exam data
                    ogretim_elemani = course_exams[0].get('ogretim_elemani', '')
                    
                    # Use ders_adi as display name
                    ders_display = ders_adi
                    
                    ws.append([date_str, time_str, ders_display, ogretim_elemani, derslikler])
                    
                    # Style data cells
                    for col in range(1, 6):
                        cell = ws.cell(row=current_row, column=col)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center' if col in [1, 2, 5] else 'left', 
                                                  vertical='center',
                                                  wrap_text=True)
                    
                    current_row += 1
            
            # Merge date cells vertically for same dates
            for date_str, start_row in date_merge_start.items():
                # Find end row for this date
                end_row = start_row
                for row in range(start_row, current_row):
                    if ws.cell(row=row, column=1).value == date_str:
                        end_row = row
                
                # Merge if multiple rows
                if end_row > start_row:
                    ws.merge_cells(f'A{start_row}:A{end_row}')
                    merged_cell = ws.cell(row=start_row, column=1)
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                    merged_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    merged_cell.font = Font(bold=True)
            
            # Set column widths
            ws.column_dimensions['A'].width = 12  # Tarih
            ws.column_dimensions['B'].width = 12  # Sınav Saati
            ws.column_dimensions['C'].width = 40  # Ders Adı
            ws.column_dimensions['D'].width = 25  # Öğretim Elemanı
            ws.column_dimensions['E'].width = 20  # Derslik
            
            # Add left border marker (colored bar like in example)
            for row in range(3, current_row):
                first_cell = ws.cell(row=row, column=1)
                # Add thick left border with color
                first_cell.border = Border(
                    left=Side(style='thick', color="3498DB"),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Save workbook
            wb.save(file_path)
            
            total_courses = len(set((s['ders_id'], s['tarih_saat']) for s in self.current_schedule))
            
            QMessageBox.information(
                self,
                "Başarılı",
                f"✅ Sınav programı Excel'e aktarıldı!\n\n"
                f"📊 {total_courses} sınav\n"
                f"📅 {len(date_merge_start)} gün\n\n"
                f"Dosya: {file_path}"
            )
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}", exc_info=True)
            QMessageBox.critical(self, "Hata", f"Excel'e aktarırken hata oluştu:\n{str(e)}")
